from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.by import By
import random
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
from playsound import playsound
import logging

driver = webdriver.Chrome(ChromeDriverManager().install())
driver.maximize_window()

actions = ActionChains(driver)


data = []
names = []


def scrape():
    rand_num = random.randint(1, 10)
    driver.get(f"http://quotes.toscrape.com/page/{rand_num}/")
    print(f"Page {rand_num} found!")
    time.sleep(3)


    quotes = driver.find_elements_by_xpath("//span[@class='text']")
    authors = driver.find_elements_by_xpath("//small[@class='author']")
    for quote in quotes:
        data.append(quote.text)

    for author in authors:
        names.append(author.text)

    print(len(data))


def Login():
    driver.get("https://twitter.com/i/flow/login")
    time.sleep(5)

    email = driver.find_element_by_xpath("//input[@name='text']").send_keys("{EnterYourEmailHere}")
    next = driver.find_element_by_xpath('//*[@id="layers"]/div/div/div/div/div/div/div[2]/div[2]/div/div/div[2]/div[2]/div[1]/div/div/div[6]').click()
    time.sleep(3)
    password = driver.find_element_by_xpath("//input[@name='password']").send_keys("{EnterYourPasswordHere}")
    login = driver.find_element_by_xpath("//div[@data-testid='LoginForm_Login_Button']")
    login.click()
    time.sleep(3)


def scroll():
    driver.execute_script("window.scrollBy(0, 2500)")

def AutoLike():
    time.sleep(3)
    for i in range(100):
        try:
            likes = driver.find_elements_by_xpath("//div[@data-testid='like']")
            if not likes:
                scroll()
            else:
                for i in range(len(likes)):
                    likes[i].click()
                    sound(2)
                    time.sleep(4)

                time.sleep(4)
        except:
            print("No likes found!")
            driver.close()

def sound(chose):
    if chose == 1:
        playsound("C:\\Users\\User\\Desktop\\class\\1st - year\\summer\\twitter_bot\\sound.mp3" , block=True)
    elif chose == 2:
        playsound("C:\\Users\\User\\Desktop\\class\\1st - year\\summer\\twitter_bot\\sheesh.mp3" , block=True)
    elif chose == 3:
        playsound("C:\\Users\\User\\Desktop\\class\\1st - year\\summer\\twitter_bot\\exploding.mp3" , block=True)



def AutoTweet_file():
    time.sleep(5)
    for i in range(5):
        tweet_box = driver.find_element_by_xpath('//*[@id="react-root"]/div/div/div[2]/main/div/div/div/div[1]/div/div[3]/div/div[2]/div[1]/div/div/div/div[2]/div[1]/div/div/div/div/div/div/div/div/div/div/label/div[1]/div/div/div/div/div/div[2]/div/div/div/div')
        rand_num = random.randint(0, 21)

        with open('C:\\Users\\User\\Desktop\\class\\1st - year\\summer\\twitter_bot\\sample.txt',encoding="utf8") as f:
            lines = f.readlines()
            tweet_box.send_keys(lines[rand_num])
        f.close()

        tweet_submit = driver.find_element_by_xpath("//div[@data-testid='tweetButtonInline']")
        tweet_submit.click()
        time.sleep(10)

def AutoTweet_Scraped():
    driver.get("https://twitter.com/home")
    time.sleep(5)
    while True:
        if not data:
            print("No more quotes to tweet!")
            sound(3)
            driver.close()
            break
        tweet_box = driver.find_element_by_xpath('//*[@id="react-root"]/div/div/div[2]/main/div/div/div/div[1]/div/div[3]/div/div[2]/div[1]/div/div/div/div[2]/div[1]/div/div/div/div/div/div/div/div/div/div/label/div[1]/div/div/div/div/div/div[2]/div/div/div/div')
        rand_num = random.randint(0, len(data) - 1)
        time.sleep(2)
        print("**************************************")
        print(f"leng of data is {len(data)}")
        print(f"random number is {rand_num} ")
        if len(data[rand_num]) > 260:
            data.pop(rand_num)
            print("Invalid tweet has been removed.")
        else:

            tweet = data[rand_num] + " --- " + names[rand_num]


            time.sleep(2)
            print(f"length of string data is {len(data[rand_num])}")
            print(data[rand_num])


            tweet_box.send_keys(tweet)
            data.pop(rand_num)


            tweet_submit = driver.find_element_by_xpath("//div[@data-testid='tweetButtonInline']")
            tweet_submit.click()
            print("success!")
            sound(1)
            time.sleep(10)


def Web_Scrapper():
    time.sleep(3)

    Search_Box = driver.find_element_by_xpath("//input[@data-testid='SearchBox_Search_Input']")
    Search_Box.send_keys("thequote")
    time.sleep(2)
    actions.send_keys(Keys.ARROW_DOWN).perform()
    actions.send_keys(Keys.ARROW_DOWN).perform()
    time.sleep(2)
    actions.send_keys(Keys.ENTER).perform()

    time.sleep(2)

    tweets = driver.find_elements_by_xpath("//div[@data-testid='tweetText']")

    time.sleep(1)
    driver.execute_script("window.scrollTo(0,300)")

    storage = load_workbook("C:\\Users\\User\\Desktop\\class\\1st - year\\summer\\twitter_bot\\\\quotes.xlsx")
    wb = Workbook()

    ws = storage.active

    ws.append("hello")
    storage.save()

    # --------- Not yet working --------- #
    # texts = driver.find_elements_by_xpath("//div[@data-testid='tweetText']")

    # try:
    #     if not texts:
    #         print("No tweets found!")
    #     else:
    #         for i in range(len(texts)):
    #             ws.append(texts[i])
    # except:
    #     print("Something went wrong!")








scrape()
Login()

# AutoTweet_file()
AutoLike()
# AutoTweet_Scraped()












